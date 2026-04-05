from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import tempfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from briarwood.agents.comparable_sales.import_csv import (
    load_active_listing_rows,
    load_comp_rows,
    merge_active_rows,
    merge_rows,
)


DEFAULT_DROP_DIR = Path.home() / "Library" / "Mobile Documents" / "com~apple~CloudDocs" / "Briarwood"
DEFAULT_MANIFEST_PATH = Path("data/comps/import_manifest.json")
DEFAULT_ARCHIVE_DIRNAME = "processed"
SUPPORTED_SUFFIXES = {".csv", ".xlsx", ".xlsm"}
ACTIVE_FILE_PATTERN = re.compile(r"^activecomps_(\d{4})(?:\..+)?$", re.IGNORECASE)
SOLD_FILE_PATTERN = re.compile(r"^soldcomps_(\d{4})(?:\..+)?$", re.IGNORECASE)


@dataclass(slots=True)
class ImportedFileResult:
    filename: str
    file_kind: str
    sold_created: int
    sold_updated: int
    active_created: int
    active_updated: int
    skipped_rows: int = 0
    error: str | None = None


def process_drop_folder(
    *,
    drop_dir: str | Path = DEFAULT_DROP_DIR,
    manifest_path: str | Path = DEFAULT_MANIFEST_PATH,
    comps_path: str | Path = "data/comps/sales_comps.json",
    active_listings_path: str | Path = "data/comps/active_listings.json",
    town: str = "Belmar",
    state: str = "NJ",
    source_name: str = "icloud weekly comp drop",
    dataset_name: str | None = None,
    as_of: str | None = None,
    archive_processed: bool = True,
) -> list[ImportedFileResult]:
    drop_dir = Path(drop_dir).expanduser()
    manifest_path = Path(manifest_path)
    archive_dir = drop_dir / DEFAULT_ARCHIVE_DIRNAME
    effective_as_of = as_of or datetime.today().strftime("%Y-%m-%d")

    manifest = _load_manifest(manifest_path)
    processed = manifest.setdefault("processed_files", {})
    results: list[ImportedFileResult] = []

    for file_path in _candidate_files(drop_dir):
        file_kind = _classify_drop_file(file_path)
        if file_kind is None:
            continue
        fingerprint = _fingerprint(file_path)
        prior = processed.get(file_path.name)
        if prior == fingerprint:
            continue

        normalized_csv = _normalize_to_csv(file_path)
        try:
            sold_created = sold_updated = active_created = active_updated = 0
            skipped_rows = 0
            error = None
            prepared_csv, skipped_rows = _prepare_csv_for_kind(normalized_csv, file_kind)
            try:
                if file_kind == "sold":
                    sold_rows = load_comp_rows(
                        prepared_csv,
                        town=town,
                        state=state,
                        source_name=source_name,
                        as_of=effective_as_of,
                    )
                    sold_created, sold_updated = merge_rows(
                        comps_path=comps_path,
                        imported_rows=sold_rows,
                        dataset_name=dataset_name,
                        as_of=effective_as_of,
                    )
                elif file_kind == "active":
                    active_rows = load_active_listing_rows(
                        prepared_csv,
                        town=town,
                        state=state,
                        source_name=source_name,
                    )
                    active_created, active_updated = merge_active_rows(
                        active_path=active_listings_path,
                        imported_rows=active_rows,
                        dataset_name=dataset_name,
                        as_of=effective_as_of,
                    )
            except Exception as exc:  # noqa: BLE001
                error = str(exc)
            finally:
                if prepared_csv != normalized_csv:
                    try:
                        prepared_csv.unlink(missing_ok=True)
                    except Exception:
                        pass
        finally:
            try:
                normalized_csv.unlink(missing_ok=True)
            except Exception:
                pass

        results.append(
            ImportedFileResult(
                filename=file_path.name,
                file_kind=file_kind,
                sold_created=sold_created,
                sold_updated=sold_updated,
                active_created=active_created,
                active_updated=active_updated,
                skipped_rows=skipped_rows,
                error=error,
            )
        )
        if error is None:
            processed[file_path.name] = fingerprint
        else:
            processed.pop(file_path.name, None)
        if archive_processed and error is None:
            archive_dir.mkdir(parents=True, exist_ok=True)
            target = archive_dir / file_path.name
            if target.exists():
                target = archive_dir / f"{file_path.stem}-{effective_as_of}{file_path.suffix}"
            shutil.move(str(file_path), str(target))

    manifest["last_run_at"] = datetime.now().isoformat(timespec="seconds")
    _write_manifest(manifest_path, manifest)
    return results


def _candidate_files(drop_dir: Path) -> Iterable[Path]:
    if not drop_dir.exists():
        return []
    return sorted(
        [
            path
            for path in drop_dir.iterdir()
            if path.is_file()
            and path.suffix.lower() in SUPPORTED_SUFFIXES
            and _classify_drop_file(path) is not None
        ],
        key=lambda path: path.name.lower(),
    )


def _classify_drop_file(path: Path) -> str | None:
    name = path.name
    if ACTIVE_FILE_PATTERN.match(name):
        return "active"
    if SOLD_FILE_PATTERN.match(name):
        return "sold"
    return None


def _normalize_to_csv(path: Path) -> Path:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return path

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        raise ValueError(f"{path.name} does not contain any worksheet rows.")

    header = [str(value).strip() if value is not None else "" for value in rows[0]]
    if not any(header):
        raise ValueError(f"{path.name} is missing a usable header row.")

    temp_file = tempfile.NamedTemporaryFile("w", suffix=".csv", delete=False, newline="", encoding="utf-8")
    with temp_file as handle:
        writer = csv.writer(handle)
        writer.writerow(header)
        for row in rows[1:]:
            writer.writerow(["" if value is None else value for value in row[: len(header)]])
    return Path(temp_file.name)


def _prepare_csv_for_kind(path: Path, file_kind: str) -> tuple[Path, int]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        fieldnames = list(reader.fieldnames or [])
        rows = list(reader)

    if not fieldnames:
        return path, 0

    cleaned_rows: list[dict[str, object]] = []
    skipped_rows = 0
    for raw in rows:
        row = {key: (value.strip() if isinstance(value, str) else value) for key, value in raw.items()}
        if not _row_has_any_values(row):
            continue
        if file_kind == "active" and _is_blank(_lookup(row, "list_price", "list price")):
            skipped_rows += 1
            continue
        if file_kind == "sold" and (_is_blank(_lookup(row, "sale_price", "sale price")) or _is_blank(_lookup(row, "sale_date", "sale date"))):
            skipped_rows += 1
            continue

        year_key = _first_present_key(row, "year_built", "year built")
        if year_key is not None:
            year_value = row.get(year_key)
            try:
                if year_value not in (None, "", "0", 0) and int(float(str(year_value))) < 1800:
                    row[year_key] = ""
                elif year_value in ("0", 0):
                    row[year_key] = ""
            except Exception:
                row[year_key] = ""

        cleaned_rows.append(row)

    if skipped_rows == 0 and cleaned_rows == rows:
        return path, 0

    temp_file = tempfile.NamedTemporaryFile("w", suffix=".csv", delete=False, newline="", encoding="utf-8")
    with temp_file as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in cleaned_rows:
            writer.writerow(row)
    return Path(temp_file.name), skipped_rows


def _row_has_any_values(row: dict[str, object]) -> bool:
    return any(not _is_blank(value) for value in row.values())


def _is_blank(value: object) -> bool:
    return value in (None, "", "N/A", "n/a", "--", "-")


def _lookup(row: dict[str, object], *keys: str) -> object:
    for key in keys:
        for existing_key, value in row.items():
            if existing_key is None:
                continue
            normalized = existing_key.strip().lower().replace("-", "_").replace(" ", "_")
            if normalized == key.replace("-", "_").replace(" ", "_"):
                return value
    return None


def _first_present_key(row: dict[str, object], *keys: str) -> str | None:
    for key in keys:
        for existing_key in row:
            if existing_key is None:
                continue
            normalized = existing_key.strip().lower().replace("-", "_").replace(" ", "_")
            if normalized == key.replace("-", "_").replace(" ", "_"):
                return existing_key
    return None


def _fingerprint(path: Path) -> dict[str, object]:
    stat = path.stat()
    return {
        "size": stat.st_size,
        "mtime_ns": stat.st_mtime_ns,
    }


def _load_manifest(path: Path) -> dict[str, object]:
    if not path.exists():
        return {"processed_files": {}}
    try:
        payload = json.loads(path.read_text())
        if isinstance(payload, dict):
            return payload
    except Exception:
        pass
    return {"processed_files": {}}


def _write_manifest(path: Path, payload: dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2) + "\n")


def main() -> int:
    parser = argparse.ArgumentParser(description="Pick up dropped CSV/XLSX comp files from an iCloud Briarwood folder and merge them into the canonical comp datasets.")
    parser.add_argument("--drop-dir", default=str(DEFAULT_DROP_DIR))
    parser.add_argument("--manifest", default=str(DEFAULT_MANIFEST_PATH))
    parser.add_argument("--comps", default="data/comps/sales_comps.json")
    parser.add_argument("--active-listings", default="data/comps/active_listings.json")
    parser.add_argument("--town", default="Belmar")
    parser.add_argument("--state", default="NJ")
    parser.add_argument("--source-name", default="icloud weekly comp drop")
    parser.add_argument("--dataset-name", default=None)
    parser.add_argument("--as-of", default=datetime.today().strftime("%Y-%m-%d"))
    parser.add_argument("--keep-files", action="store_true", help="Do not move processed files into the processed/ archive.")
    args = parser.parse_args()

    results = process_drop_folder(
        drop_dir=args.drop_dir,
        manifest_path=args.manifest,
        comps_path=args.comps,
        active_listings_path=args.active_listings,
        town=args.town,
        state=args.state,
        source_name=args.source_name,
        dataset_name=args.dataset_name,
        as_of=args.as_of,
        archive_processed=not args.keep_files,
    )

    if not results:
        print(f"No new comp files found in {args.drop_dir}")
        return 0

    for result in results:
        if result.error:
            print(f"{result.filename} [{result.file_kind}]: FAILED - {result.error}")
            continue
        print(
            f"{result.filename} [{result.file_kind}]: sold {result.sold_created} created / {result.sold_updated} updated, "
            f"active {result.active_created} created / {result.active_updated} updated"
            + (f", skipped {result.skipped_rows} invalid rows" if result.skipped_rows else "")
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
