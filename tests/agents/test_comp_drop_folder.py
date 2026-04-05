import csv
import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from audit_scripts.pickup_comp_drop_folder import process_drop_folder


class CompDropFolderWorkflowTests(unittest.TestCase):
    def test_process_drop_folder_routes_sold_file_to_sales_dataset(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            drop_dir = base / "drop"
            drop_dir.mkdir()
            manifest = base / "manifest.json"
            comps = base / "sales.json"
            active = base / "active.json"
            comps.write_text(json.dumps({"metadata": {}, "sales": []}))
            active.write_text(json.dumps({"metadata": {}, "listings": []}))

            workbook_path = drop_dir / "soldcomps_0405.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["address", "sale_price", "sale_date", "sqft", "beds", "baths", "lot_size", "year_built", "verification_status", "status"])
            ws.append(["304 8th Ave", 810000, "2026-03-15", 1400, 3, 2, 0.11, 1950, "manual", "sold"])
            wb.save(workbook_path)

            results = process_drop_folder(
                drop_dir=drop_dir,
                manifest_path=manifest,
                comps_path=comps,
                active_listings_path=active,
                town="Belmar",
                state="NJ",
                as_of="2026-04-05",
            )

            self.assertEqual(len(results), 1)
            self.assertEqual(results[0].file_kind, "sold")
            self.assertFalse(workbook_path.exists())
            self.assertTrue((drop_dir / "processed" / "soldcomps_0405.xlsx").exists())

            sold_payload = json.loads(comps.read_text())
            active_payload = json.loads(active.read_text())
            self.assertEqual(len(sold_payload["sales"]), 1)
            self.assertEqual(len(active_payload["listings"]), 0)
            self.assertEqual(sold_payload["metadata"]["as_of"], "2026-04-05")

    def test_process_drop_folder_routes_active_file_to_active_dataset(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            drop_dir = base / "drop"
            drop_dir.mkdir()
            manifest = base / "manifest.json"
            comps = base / "sales.json"
            active = base / "active.json"
            comps.write_text(json.dumps({"metadata": {}, "sales": []}))
            active.write_text(json.dumps({"metadata": {}, "listings": []}))

            workbook_path = drop_dir / "activecomps_0405.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["address", "list_price", "sqft", "beds", "baths", "lot_size", "year_built", "verification_status", "status"])
            ws.append(["1223 Briarwood Rd", 829000, 1820, 3, 2, 0.17, 1952, "manual", "active"])
            wb.save(workbook_path)

            results = process_drop_folder(
                drop_dir=drop_dir,
                manifest_path=manifest,
                comps_path=comps,
                active_listings_path=active,
                town="Belmar",
                state="NJ",
                as_of="2026-04-05",
            )

            self.assertEqual(len(results), 1)
            self.assertEqual(results[0].file_kind, "active")
            sold_payload = json.loads(comps.read_text())
            active_payload = json.loads(active.read_text())
            self.assertEqual(len(sold_payload["sales"]), 0)
            self.assertEqual(len(active_payload["listings"]), 1)

    def test_process_drop_folder_skips_invalid_active_rows_but_loads_valid_ones(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            drop_dir = base / "drop"
            drop_dir.mkdir()
            manifest = base / "manifest.json"
            comps = base / "sales.json"
            active = base / "active.json"
            comps.write_text(json.dumps({"metadata": {}, "sales": []}))
            active.write_text(json.dumps({"metadata": {}, "listings": []}))

            workbook_path = drop_dir / "activecomps_0405.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["address", "list_price", "sqft", "beds", "baths", "lot_size", "year_built", "verification_status", "status"])
            ws.append(["Good Row", 829000, 1820, 3, 2, 0.17, 1952, "manual", "active"])
            ws.append(["Missing Price", "", 1600, 3, 2, 0.10, 1948, "manual", "active"])
            ws.append(["Zero Year", 745000, 1500, 3, 2, 0.11, 0, "manual", "active"])
            wb.save(workbook_path)

            results = process_drop_folder(
                drop_dir=drop_dir,
                manifest_path=manifest,
                comps_path=comps,
                active_listings_path=active,
                town="Belmar",
                state="NJ",
                as_of="2026-04-05",
            )

            self.assertEqual(len(results), 1)
            self.assertEqual(results[0].file_kind, "active")
            self.assertEqual(results[0].skipped_rows, 1)
            self.assertIsNone(results[0].error)
            active_payload = json.loads(active.read_text())
            self.assertEqual(len(active_payload["listings"]), 2)
            zero_year_row = next(row for row in active_payload["listings"] if row["address"] == "Zero Year")
            self.assertNotIn("year_built", zero_year_row)

    def test_process_drop_folder_skips_already_processed_files(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            drop_dir = base / "drop"
            drop_dir.mkdir()
            manifest = base / "manifest.json"
            comps = base / "sales.json"
            active = base / "active.json"
            comps.write_text(json.dumps({"metadata": {}, "sales": []}))
            active.write_text(json.dumps({"metadata": {}, "listings": []}))

            csv_path = drop_dir / "soldcomps_0405.csv"
            with csv_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=["address", "sale_price", "sale_date", "sqft", "beds", "baths", "lot_size", "year_built", "verification_status", "status"],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "address": "304 8th Ave",
                        "sale_price": "810000",
                        "sale_date": "2026-03-15",
                        "sqft": "1400",
                        "beds": "3",
                        "baths": "2",
                        "lot_size": "0.11",
                        "year_built": "1950",
                        "verification_status": "manual",
                        "status": "sold",
                    }
                )

            first = process_drop_folder(
                drop_dir=drop_dir,
                manifest_path=manifest,
                comps_path=comps,
                active_listings_path=active,
                town="Belmar",
                state="NJ",
                as_of="2026-04-05",
                archive_processed=False,
            )
            second = process_drop_folder(
                drop_dir=drop_dir,
                manifest_path=manifest,
                comps_path=comps,
                active_listings_path=active,
                town="Belmar",
                state="NJ",
                as_of="2026-04-05",
                archive_processed=False,
            )

            self.assertEqual(len(first), 1)
            self.assertEqual(second, [])
            payload = json.loads(comps.read_text())
            self.assertEqual(len(payload["sales"]), 1)


if __name__ == "__main__":
    unittest.main()
